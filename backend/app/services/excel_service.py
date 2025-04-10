import pandas as pd
import openpyxl
from typing import Dict, Any, List, Tuple, Optional
import os
import tempfile

class ExcelService:
    """
    Service for processing Excel files, specifically for extracting data from
    safety plan and risk assessment sheets.
    """
    
    def __init__(self, file_path: Optional[str] = None, file_content: Optional[bytes] = None):
        """
        Initialize the Excel processor with either a file path or file content.
        
        Args:
            file_path: Path to the Excel file
            file_content: Content of the Excel file as bytes
        """
        if file_path:
            self.file_path = file_path
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        elif file_content:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp_file:
                tmp_file.write(file_content)
                self.file_path = tmp_file.name
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        else:
            raise ValueError("Either file_path or file_content must be provided")
    
    def extract_data(self) -> Dict[str, Any]:
        """
        Extract data from the Excel file according to the specified cell ranges.
        
        Returns:
            Dict containing the extracted data from safety plan and risk assessment sheets
        """
        result = {}
        
        safety_plan_data = self._extract_safety_plan_data()
        if safety_plan_data:
            result["safety_plan"] = safety_plan_data
        
        risk_assessment_data = self._extract_risk_assessment_data()
        if risk_assessment_data:
            result["risk_assessment"] = risk_assessment_data
        
        return result
    
    def _extract_safety_plan_data(self) -> Dict[str, Any]:
        """
        Extract data from the safety plan sheet.
        
        Returns:
            Dict containing the extracted data from the safety plan sheet
        """
        sheet_name = self._find_sheet_name(["安全施工計画書", "安全施工計画", "施工計画書"])
        
        if not sheet_name:
            return {}
        
        sheet = self.workbook[sheet_name]
        
        columns = ["D", "AS", "BK"]
        row_ranges = [(34, 43), (54, 73), (81, 100), (108, 127)]
        
        data = {}
        
        for col in columns:
            col_data = {}
            for start_row, end_row in row_ranges:
                for row in range(start_row, end_row + 1):
                    cell_value = sheet[f"{col}{row}"].value
                    if cell_value:  # Only include non-empty cells
                        col_data[f"Row_{row}"] = str(cell_value)
            
            if col_data:  # Only include columns with data
                data[f"Column_{col}"] = col_data
        
        return data
    
    def _extract_risk_assessment_data(self) -> Dict[str, Any]:
        """
        Extract data from the risk assessment sheet.
        
        Returns:
            Dict containing the extracted data from the risk assessment sheet
        """
        sheet_name = self._find_sheet_name(["リスクアセスメント", "リスク評価", "リスク分析"])
        
        if not sheet_name:
            return {}
        
        sheet = self.workbook[sheet_name]
        
        columns = ["C", "O", "Y", "AI", "AS", "CC", "AV", "CF", "BE", "CL"]
        
        rows = [
            34, 37, 40, 48, 51, 54, 57, 60, 63, 66, 69, 72, 75, 78, 81, 84, 87, 
            96, 99, 102, 105, 108, 111, 114, 117, 120, 123, 126, 129, 132, 135,
            144, 147, 150, 153, 156, 159, 162, 165, 168, 171, 174, 177, 180, 183,
            192, 195, 198, 201, 204, 207, 210, 213, 216, 219, 222, 225, 228, 231,
            240, 243, 246, 249, 252, 255, 258, 261, 264, 267, 270, 273, 276, 279,
            288, 291, 294, 297, 300, 303, 306, 309, 312, 315, 318, 321, 324, 327,
            336, 339, 342, 345, 348, 351, 354, 357, 360, 363, 366, 369, 372, 375,
            384, 387, 390, 393, 396, 399, 402, 405, 408, 411, 414, 417, 420, 423,
            432, 435, 438, 441, 444, 447, 450, 453, 456, 459, 462, 465, 468, 471,
            480, 483, 486, 489, 492, 495, 498, 501, 504, 507, 510, 513, 516, 519,
            528, 531, 534, 537, 540, 543, 546, 549, 552, 555, 558, 561, 564, 567
        ]
        
        data = {}
        
        for col in columns:
            col_data = {}
            for row in rows:
                cell_value = sheet[f"{col}{row}"].value
                if cell_value:  # Only include non-empty cells
                    col_data[f"Row_{row}"] = str(cell_value)
            
            if col_data:  # Only include columns with data
                data[f"Column_{col}"] = col_data
        
        return data
    
    def _find_sheet_name(self, possible_names: List[str]) -> str:
        """
        Find a sheet name that matches one of the possible names.
        
        Args:
            possible_names: List of possible sheet names
            
        Returns:
            The name of the sheet if found, otherwise an empty string
        """
        for sheet_name in self.workbook.sheetnames:
            for possible_name in possible_names:
                if possible_name in sheet_name:
                    return sheet_name
        
        return ""
    
    def get_fixed_info(self) -> Dict[str, str]:
        """
        Get the fixed information from the Excel file.
        
        Returns:
            Dict containing the fixed information (project name, location, period, workers)
        """
        sheet_name = self._find_sheet_name(["安全施工計画書", "安全施工計画", "施工計画書"])
        
        if not sheet_name:
            return {
                "project_name": "",
                "location": "",
                "period": "",
                "workers": ""
            }
        
        sheet = self.workbook[sheet_name]
        
        project_name = sheet["I9"].value or ""
        location = sheet["I11"].value or ""
        period = sheet["I13"].value or ""
        workers = sheet["Q15"].value or ""
        
        return {
            "project_name": str(project_name),
            "location": str(location),
            "period": str(period),
            "workers": str(workers)
        }
    
    def cleanup(self):
        """
        Clean up temporary files.
        """
        if hasattr(self, 'file_path') and os.path.exists(self.file_path) and self.file_path.startswith(tempfile.gettempdir()):
            os.unlink(self.file_path)
